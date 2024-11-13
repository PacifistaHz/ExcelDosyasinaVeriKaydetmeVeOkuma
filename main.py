from openpyxl import Workbook, load_workbook
wb = load_workbook('Veriler.xlsx')
ws = wb["Sayfa1"]

print(ws.cell(row=1, column=1).value)

print(ws["A3"].value)

for satir in ws["A1":"B6"]:
    for hucra in satir:
        print(str(hucra.value) + "\t", end="")
    print()

for satir in range(1,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print("|" + str(ws.cell(row=satir, column=sutun).value) + "|", end="")
    print()
wb.close()